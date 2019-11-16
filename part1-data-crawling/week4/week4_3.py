import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active     # 활성화된 sheet

# 위치명을 파악하고 있어야 한다.
sheet['A1'] = 'Hello World!'

# 그럴 필요 없이 row와 column을 위치 지정이 가능하다.
sheet.cell(row=3, column=4).value = 'Goodbye!'

for i in range(1, 10):
    sheet.cell(row=i, column=5).value = i

# append에 리스트를 넣으면 엑셀에 추가
# 가장 마지막 줄에 데이터를 알아서 넣어줌
# 정말로 위치를 지정하지 않아도 된다. -> 자동화 가능
test = ['채널', '재생 수', '좋아요 수']
sheet.append(test)

wb.save('test.xlsx')