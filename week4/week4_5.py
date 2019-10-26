import requests
from bs4 import BeautifulSoup
import openpyxl

MAX = 20
KEYWORD = input('검색어를 입력해주세요 : ')

# wb = openpyxl.Workbook()                          # 엑셀 파일 생성
# wb = openpyxl.load_workbook('navernews.xlsx')     # 엑셀 파일 불러오기
try:
    wb = openpyxl.load_workbook('navernews.xlsx')
    sheet = wb.active
except:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(['검색어', '기사', '언론사'])

for page in range(1, MAX, 10):
    raw = requests.get('https://search.naver.com/search.naver?where=news&sm=tab_jum&query='+KEYWORD+'&start='+str(page), headers={'User-Agent':'Mozilla/5.0'})
    code = BeautifulSoup(raw.text, 'html.parser')

    container = code.select('ul.type01 > li')

    for cont in container:
        # 기사명
        title = cont.select_one('a._sp_each_title').text.strip()
        # 언론사명
        writer = cont.select_one('span._sp_each_source').text.strip()

        print(title, writer)
        sheet.append([KEYWORD, title, writer])

wb.save('navernews.xlsx')